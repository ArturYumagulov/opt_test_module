import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from db import Results


engine = create_engine("sqlite:///products_28,10.db")
Session = sessionmaker(bind=engine)
s = Session()

data = s.query(Results).all()
result = []
for i in data:
    result.append([i.test_date, i.group, i.name, i.result_name, i.article, i.result_article, i.catalog_number,
                   i.result_catalog_num])


def write_xls(lst, path):
    wb = openpyxl.Workbook()
    wb.create_sheet(title="Лист", index=0)
    sheet = wb["Лист"]
    for i in range(len(lst)):
        for j in lst[i]:
            value = str(j)
            cell = sheet.cell(row=i + 1, column=(lst[i].index(j)) + 1)
            cell.value = value
    wb.save(path)


write_xls(result, "result.xlsx")